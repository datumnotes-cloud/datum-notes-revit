# -*- coding: utf-8 -*-
"""
Excel Template & Import Module for Datum Notes Redline pyRevit Extension

This module provides functionality to:
1. Generate an Excel template with strict formatting and AI instructions
2. Read and parse completed Excel files
3. Validate Excel data with detailed error reporting
"""

from __future__ import print_function

import os
import datetime
import json

try:
    # Try to import openpyxl for Excel support
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, DEFAULT_FONT
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ============================================================================
# CONSTANTS
# ============================================================================

EXCEL_COLUMNS = {
    "room": {"col": "A", "width": 18, "required": True},
    "note": {"col": "B", "width": 45, "required": True},
    "category": {"col": "C", "width": 16, "required": True},
    "assignedTo": {"col": "D", "width": 18, "required": False},
    "dueDate": {"col": "E", "width": 14, "required": False},
}

VALID_CATEGORIES = ["Action Item", "Decision", "Question", "Observation"]

# Color scheme for headers
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
INSTRUCTION_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
INSTRUCTION_FONT = Font(size=10, italic=True, color="404040")
BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


# ============================================================================
# EXCEL TEMPLATE GENERATION
# ============================================================================

def create_excel_template(project_name, rooms_list=None):
    """
    Create an Excel workbook with instructions and template data.
    
    Args:
        project_name (str): The Revit project name
        rooms_list (list): Optional list of room names for auto-complete suggestions
    
    Returns:
        Workbook: openpyxl Workbook object
    """
    if not OPENPYXL_AVAILABLE:
        return None
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Datum Import"
    
    # ===== TITLE & PROJECT INFO =====
    ws['A1'] = "DATUM NOTES - IMPORT TEMPLATE"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    
    ws['A2'] = "Project:"
    ws['B2'] = project_name
    ws['A2'].font = Font(bold=True)
    ws['B2'].font = Font(size=11)
    
    ws['A3'] = "Date:"
    ws['B3'] = datetime.datetime.now().strftime("%Y-%m-%d")
    ws['A3'].font = Font(bold=True)
    ws['B3'].font = Font(size=11)
    
    # ===== INSTRUCTIONS SECTION =====
    ws['A5'] = "INSTRUCTIONS FOR AI ASSISTANT"
    ws['A5'].font = Font(bold=True, size=11, underline="single")
    ws.merge_cells('A5:E5')
    
    instructions = [
        "",
        "1. ROOM IDENTIFICATION:",
        "   - If a note relates to a specific room, enter the room name/number in column A",
        "   - If not sure which room, enter: UNASSIGNED",
        "   - Be EXACT: spaces, hyphens, and case matter",
        "",
        "2. NOTE TEXT (Column B):",
        "   - Enter a clear, concise description of the action, decision, question, or observation",
        "   - Keep to one sentence or short paragraph",
        "   - Do NOT include extra formatting or special characters (except basic punctuation)",
        "",
        "3. CATEGORY (Column C) - MUST be EXACTLY one of these:",
        "   - 'Action Item' (for tasks/to-dos)",
        "   - 'Decision' (for resolved choices)",
        "   - 'Question' (for open items needing clarification)",
        "   - 'Observation' (for notes/comments)",
        "",
        "4. ASSIGNED TO (Column D) - OPTIONAL:",
        "   - Enter person's name if task is assigned",
        "   - Leave blank if not assigned",
        "",
        "5. DUE DATE (Column E) - OPTIONAL:",
        "   - Format: YYYY-MM-DD (e.g., 2025-03-20)",
        "   - Only for Action Items; ignore for other categories",
        "",
        "6. DO NOT:",
        "   - Delete row headers",
        "   - Add or remove columns",
        "   - Merge cells",
        "   - Leave required fields (Room, Note, Category) empty",
        "   - Add extra sheets (all data must be on this sheet)",
    ]
    
    row = 6
    for instruction in instructions:
        ws[f'A{row}'] = instruction
        ws[f'A{row}'].font = INSTRUCTION_FONT
        ws.merge_cells(f'A{row}:E{row}')
        if instruction.strip().startswith(("1.", "2.", "3.", "4.", "5.", "6.", "DO NOT:")):
            ws[f'A{row}'].font = Font(bold=True, size=10, color="404040")
        row += 1
    
    # ===== COLUMN HEADERS =====
    header_row = row + 1
    headers = [
        ("Room Name / Number", "A"),
        ("Note Description", "B"),
        ("Category", "C"),
        ("Assigned To", "D"),
        ("Due Date (YYYY-MM-DD)", "E"),
    ]
    
    for header_text, col in headers:
        cell = ws[f'{col}{header_row}']
        cell.value = header_text
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN
    
    ws.row_dimensions[header_row].height = 30
    
    # ===== EXAMPLE ROW =====
    example_row = header_row + 1
    example_data = [
        "Conference Room B",
        "Review HVAC design with MEP",
        "Decision",
        "MEP",
        "2025-03-25"
    ]
    
    for idx, (data, col) in enumerate(zip(example_data, ["A", "B", "C", "D", "E"])):
        cell = ws[f'{col}{example_row}']
        cell.value = data
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        cell.border = BORDER_THIN
        cell.font = Font(size=10, italic=True, color="808080")
    
    ws.row_dimensions[example_row].height = 20
    
    # ===== EMPTY DATA ROWS (20 rows for user to fill) =====
    for data_row in range(example_row + 1, example_row + 21):
        for col in ["A", "B", "C", "D", "E"]:
            cell = ws[f'{col}{data_row}']
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[data_row].height = 25
    
    # ===== DROPDOWN SUGGESTIONS (Optional) =====
    if rooms_list and len(rooms_list) > 0:
        # Add a reference sheet with room names
        ref_sheet = wb.create_sheet("Rooms Reference")
        ref_sheet.column_dimensions['A'].width = 25
        ref_sheet['A1'] = "Available Rooms"
        ref_sheet['A1'].font = Font(bold=True)
        
        for idx, room in enumerate(rooms_list[:100], start=2):  # Limit to 100
            ref_sheet[f'A{idx}'] = room
    
    # ===== SET COLUMN WIDTHS =====
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    
    return wb


def save_excel_template(workbook, filepath):
    """Save Excel workbook to file."""
    if not OPENPYXL_AVAILABLE:
        return False
    
    try:
        workbook.save(filepath)
        return True
    except Exception as ex:
        print("Error saving Excel template: %s" % str(ex))
        return False


# ============================================================================
# EXCEL FILE PARSING & VALIDATION
# ============================================================================

def validate_excel_file(filepath):
    """
    Validate that the Excel file has proper structure.
    
    Returns:
        tuple: (is_valid, error_message)
    """
    if not OPENPYXL_AVAILABLE:
        return False, "openpyxl library not available. Please use text-based import."
    
    try:
        wb = load_workbook(filepath)
    except Exception as ex:
        return False, "Cannot open Excel file: %s" % str(ex)
    
    if "Datum Import" not in wb.sheetnames and len(wb.sheetnames) == 0:
        return False, "No valid sheet found. Expected 'Datum Import' sheet."
    
    ws = wb["Datum Import"] if "Datum Import" in wb.sheetnames else wb.active
    
    # Check for header row (should be around row 35-36 after instructions)
    header_row = None
    for row_num in range(1, min(50, ws.max_row + 1)):
        cell_a = ws[f'A{row_num}'].value
        if cell_a and str(cell_a).strip().lower().startswith("room"):
            header_row = row_num
            break
    
    if header_row is None:
        return False, "Could not find header row. Template may be malformed."
    
    return True, ""


def parse_excel_file(filepath, normalize_fn, match_room_fn, unassigned_bucket_fn):
    """
    Parse completed Excel file and extract note data.
    
    Args:
        filepath (str): Path to Excel file
        normalize_fn (callable): Function to normalize category
        match_room_fn (callable): Function to match room from text
        unassigned_bucket_fn (callable): Function to get unassigned room bucket
    
    Returns:
        tuple: (parsed_items, error_message)
    """
    if not OPENPYXL_AVAILABLE:
        return [], "openpyxl library not available."
    
    try:
        wb = load_workbook(filepath)
    except Exception as ex:
        return [], "Cannot open Excel file: %s" % str(ex)
    
    ws = wb["Datum Import"] if "Datum Import" in wb.sheetnames else wb.active
    
    # Find header row
    header_row = None
    for row_num in range(1, min(50, ws.max_row + 1)):
        cell_a = ws[f'A{row_num}'].value
        if cell_a and str(cell_a).strip().lower().startswith("room"):
            header_row = row_num
            break
    
    if header_row is None:
        return [], "Could not find header row in Excel file."
    
    # Parse data rows (starting from header_row + 1)
    parsed = []
    errors = []
    data_row = header_row + 1
    row_count = 0
    
    while data_row <= ws.max_row:
        # Read cells
        room_cell = ws[f'A{data_row}'].value
        note_cell = ws[f'B{data_row}'].value
        category_cell = ws[f'C{data_row}'].value
        assigned_cell = ws[f'D{data_row}'].value
        due_cell = ws[f'E{data_row}'].value
        
        room_val = str(room_cell).strip() if room_cell else ""
        note_val = str(note_cell).strip() if note_cell else ""
        category_val = str(category_cell).strip() if category_cell else ""
        assigned_val = str(assigned_cell).strip() if assigned_cell else ""
        due_val = str(due_cell).strip() if due_cell else ""
        
        # Skip empty rows
        if not room_val and not note_val and not category_val:
            data_row += 1
            continue
        
        row_count += 1
        
        # Validate required fields
        if not room_val:
            errors.append("Row %d: Room is required" % data_row)
            data_row += 1
            continue
        
        if not note_val:
            errors.append("Row %d: Note description is required" % data_row)
            data_row += 1
            continue
        
        if not category_val:
            errors.append("Row %d: Category is required (Action Item, Decision, Question, or Observation)" % data_row)
            data_row += 1
            continue
        
        # Validate category
        if category_val not in VALID_CATEGORIES:
            errors.append("Row %d: Invalid category '%s'. Must be: %s" % (
                data_row, category_val, ", ".join(VALID_CATEGORIES)
            ))
            data_row += 1
            continue
        
        # Validate date format if provided
        if due_val:
            if not is_valid_date(due_val):
                errors.append("Row %d: Invalid date format '%s'. Use YYYY-MM-DD" % (data_row, due_val))
                data_row += 1
                continue
        
        # Build parsed item
        parsed.append({
            "room": room_val,
            "text": note_val,
            "category": category_val,
            "assignedTo": assigned_val,
            "dueDate": due_val,
            "isUnassigned": room_val.upper() == "UNASSIGNED"
        })
        
        data_row += 1
    
    if errors and len(parsed) == 0:
        error_msg = "Validation errors:\n" + "\n".join(errors[:5])
        if len(errors) > 5:
            error_msg += "\n... and %d more errors" % (len(errors) - 5)
        return [], error_msg
    
    return parsed, ""


def is_valid_date(date_string):
    """Check if string is valid YYYY-MM-DD format."""
    try:
        datetime.datetime.strptime(str(date_string).strip(), "%Y-%m-%d")
        return True
    except ValueError:
        return False


# ============================================================================
# STATUS CHECK
# ============================================================================

def excel_available():
    """Check if Excel support is available."""
    return OPENPYXL_AVAILABLE
